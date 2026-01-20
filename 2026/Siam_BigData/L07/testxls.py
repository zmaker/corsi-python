from openpyxl import load_workbook

wb = load_workbook("dati.xlsx")

#elenco i fogli
print(wb.sheetnames)
fogli = wb.sheetnames
print(fogli[0])
foglio_dati = fogli[0]

#seleziono il foglio

#ws = wb["Vendite"]
ws = wb[foglio_dati]

#leggo una cella
valore = ws["A2"].value
print("A2", valore)
#leggo con indici
valore = ws.cell(row=3, column=2).value
print("3,2", valore)

#Leggere una riga intera.
riga = []
for cella in ws[1]:      # riga 1
    riga.append(cella.value)

print(riga)

# Leggere tutta la tabella riga per riga
# values_only=True restituisce direttamente i valori, non gli oggetti Cell.
# parto dalla riga 2
for row in ws.iter_rows(min_row=2, values_only=True):
    print(row)

# Leggere una colonna.
colonna = []
for cella in ws["B"]:
    colonna.append(cella.value)
print(colonna)

#modifico una cella
ws["C3"] = 99

#salvare il file.
wb.save("dati_modificati.xlsx")



