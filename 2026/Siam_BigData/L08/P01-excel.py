from openpyxl import load_workbook

wb = load_workbook("./dati/caseroma.xlsx")
print(wb.sheetnames)
pages = wb.sheetnames
#seleziono il foglio di lavoro
ws = wb[pages[0]]

valore = ws["B2"].value
print(valore)

#leggo una riga
riga = []
for cella in ws[1]:
    riga.append(cella.value)
print(riga)

#modifico una cella
ws["D2"] = 99
ws["D2"].number_format = "0"

wb.save("./dati/out.xlsx")
